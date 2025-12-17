from __future__ import annotations

import time
import io
from datetime import date, timedelta
from typing import Any, Optional

import duckdb
import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from supabase_link import get_supabase_client
from aws_duckdb import get_duckdb_connection

try:
    from st_aggrid import AgGrid  # type: ignore
    from st_aggrid.grid_options_builder import GridOptionsBuilder  # type: ignore
    from st_aggrid.shared import GridUpdateMode  # type: ignore
    _HAS_AGGRID = True
except Exception:
    AgGrid = None  # type: ignore
    GridOptionsBuilder = None  # type: ignore
    GridUpdateMode = None  # type: ignore
    _HAS_AGGRID = False


# -----------------------------
# Owner-configurable reasons list
# -----------------------------

REASONS: list[str] = sorted([
    "Disconnected String","SCB Fire","HT Panel Fire","Transformer Fire","Cable Failure","Clipping Loss","Deration Loss","Degrdation Loss","Shading Loss","Bypass Diode Failure","Bypass Diode Burnt","Soiling Loss","Passing Clouds","Reactive Power Loss","Fuse Failure","IGBT Failure","AC Breaker Issue","DC Breaker Issue","Module Damage","Temperature Loss","RISO Fault","MPPT Malfunction","Efficiency Loss","Ground Fault","Module Mismatch Loss","Array Misalignment","Tracker Failure","Inverter Fan Issue","Bifacial Factor Loss","Power Limitation","AC Current Imbalance","Vegetation Growth","Low Irradiation","Manual Trip/Stop","Trip","Overvoltage","Overcurrent","LVRT/HVRT","Land Undulation","DC Loading Pending","Inverter Card Issue","Inverter Software Fault",
    "Inverter Fire",
    "Connector Burn",
    "Grid Issue",
    "Curtailment","Unknown/No Issue","Capacitor Failure/Issue","Rainy Weather","Bad Weather","Late Wakeup","Theft","Cable Failure/Puncture","Rodent Issue","Heating Issue","Hotspot Formation","Outgoing Trip","Load Shifting","LT Panel Trip","Isolator Issue","Breaker Issue","Transposition Loss","Design Loss","IGBT Temperature Issue","Communication Issue","Bad Data","GFDI Protection Operated","Half String","Hardware Fault","Phase to Phase Fault","Phase to Ground Fault","Over Temperature Trip","Reactor Failure","Hardware Failure"
])


# -----------------------------
# DuckDB helpers (SYD source)
# -----------------------------


def _connect_ro(db_path: str) -> duckdb.DuckDBPyConnection:
    return get_duckdb_connection(db_local=db_path)


@st.cache_data(show_spinner=False)
def list_sites_from_syd(db_path: str) -> list[str]:
    con = _connect_ro(db_path)
    try:
        rows = con.execute("select distinct site_name from syd order by site_name").fetchall()
        return [r[0] for r in rows]
    finally:
        con.close()


@st.cache_data(show_spinner=False)
def date_bounds_from_syd(db_path: str) -> tuple[Optional[date], Optional[date]]:
    con = _connect_ro(db_path)
    try:
        row = con.execute("select min(date) as dmin, max(date) as dmax from syd").fetchone()
        if not row:
            return None, None
        return row[0], row[1]
    finally:
        con.close()


@st.cache_data(show_spinner=False)
def equipment_deviation_candidates(
    db_path: str,
    *,
    site_name: str,
    start_date: date,
    end_date: date,
    threshold: float,
) -> pd.DataFrame:
    """
    Returns equipment candidates for dropdown with their deviation (%):
    - Single day: deviation = syd_percent*100 for that date
    - Range: deviation = avg(syd_percent*100) over the range
    Only returns equipment with deviation < threshold.
    """
    con = _connect_ro(db_path)
    try:
        if start_date == end_date:
            return con.execute(
                """
                select
                  equipment_name,
                  syd_percent * 100.0 as syd_dev_pct
                from syd
                where site_name = ?
                  and date = ?
                  and (syd_percent * 100.0) < ?
                order by syd_dev_pct asc, equipment_name asc
                """,
                [site_name, start_date, float(threshold)],
            ).fetchdf()

        return con.execute(
            """
            select
              equipment_name,
              avg(syd_percent * 100.0) as syd_dev_pct
            from syd
            where site_name = ?
              and date between ? and ?
            group by 1
            having avg(syd_percent * 100.0) < ?
            order by syd_dev_pct asc, equipment_name asc
            """,
            [site_name, start_date, end_date, float(threshold)],
        ).fetchdf()
    finally:
        con.close()


@st.cache_data(show_spinner=False)
def fetch_download_candidates(
    db_path: str,
    *,
    site_name: str,
    start_date: date,
    end_date: date,
    threshold: float
) -> pd.DataFrame:
    """
    Fetch candidates for Excel download.
    - Single Day: Actual SYD
    - Date Range: Median SYD
    """
    con = _connect_ro(db_path)
    try:
        # Single Date: Use actual value
        if start_date == end_date:
            return con.execute(
                """
                select
                  site_name,
                  equipment_name,
                  (syd_percent * 100.0) as deviation
                from syd
                where site_name = ?
                  and date = ?
                  and (syd_percent * 100.0) < ?
                order by equipment_name
                """,
                [site_name, start_date, float(threshold)]
            ).fetchdf()
        
        # Date Range: Use MEDIAN
        return con.execute(
            """
            select
              site_name,
              equipment_name,
              median(syd_percent * 100.0) as deviation
            from syd
            where site_name = ?
              and date between ? and ?
            group by 1, 2
            having median(syd_percent * 100.0) < ?
            order by equipment_name
            """,
            [site_name, start_date, end_date, float(threshold)]
        ).fetchdf()
    finally:
        con.close()


# -----------------------------
# Supabase helpers
# -----------------------------


def _to_iso(d: date) -> str:
    return d.isoformat()


@st.cache_data(show_spinner=False)
def fetch_comments_live(
    *,
    site_name: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    equipment_name: Optional[str] = None,
    limit: int = 500,
) -> pd.DataFrame:
    sb = get_supabase_client(prefer_service_role=True)
    q = sb.table("zelestra_comments").select("*").order("created_at", desc=True).limit(int(limit))

    if site_name:
        q = q.eq("site_name", site_name)

    if start_date and end_date:
        q = q.lte("start_date", end_date.isoformat()).gte("end_date", start_date.isoformat())
    elif start_date and not end_date:
        q = q.gte("end_date", start_date.isoformat())
    elif end_date and not start_date:
        q = q.lte("start_date", end_date.isoformat())

    if equipment_name:
        try:
            q = q.contains("equipment_names", [equipment_name])
        except Exception:
            pass

    resp = q.execute()
    rows = resp.data or []
    return pd.DataFrame(rows)


def _clear_comments_cache() -> None:
    pass


def insert_comment(payload: dict[str, Any]) -> None:
    # Add created_by from session state if available
    user_info = st.session_state.get("user_info", {})
    username = user_info.get("username")
    if username:
        payload["created_by"] = username
    
    sb = get_supabase_client(prefer_service_role=True)
    sb.table("zelestra_comments").insert(payload).execute()


def insert_bulk_comments(payloads: list[dict[str, Any]]) -> None:
    """Bulk insert with fallback for integer schema mismatch (Fix for 22P02 error)"""
    if not payloads:
        return
    
    # Add created_by from session state if available
    user_info = st.session_state.get("user_info", {})
    username = user_info.get("username")
    if username:
        for p in payloads:
            p["created_by"] = username
    
    sb = get_supabase_client(prefer_service_role=True)
    try:
        # Try inserting exactly as provided (likely floats)
        sb.table("zelestra_comments").insert(payloads).execute()
    except Exception as e:
        # Check if error is specifically about integer format "22P02"
        err_msg = str(e).lower()
        if "22p02" in err_msg or "invalid input syntax for type integer" in err_msg:
            # Fallback: Round all deviations to nearest integer and retry
            for p in payloads:
                if "deviation" in p:
                    p["deviation"] = int(round(p["deviation"]))
            sb.table("zelestra_comments").insert(payloads).execute()
        else:
            # Re-raise if it's a different error
            raise e


def update_comment(comment_id: Any, payload: dict[str, Any]) -> None:
    sb = get_supabase_client(prefer_service_role=True)
    sb.table("zelestra_comments").update(payload).eq("id", comment_id).execute()


def _stringify_error(e: Exception) -> str:
    try:
        return str(e)
    except Exception:
        return repr(e)


def _needs_int_deviation(err: str) -> bool:
    s = err.lower()
    return ("22p02" in s and "integer" in s) or ("invalid input syntax for type integer" in s)


def _write_comment_with_fallback(*, comment_id: Any | None, payload: dict[str, Any], deviation_value: float) -> None:
    try:
        if comment_id is None:
            insert_comment(payload)
        else:
            update_comment(comment_id, payload)
        return
    except Exception as e:
        err = _stringify_error(e)
        if _needs_int_deviation(err):
            payload2 = dict(payload)
            payload2["deviation"] = int(round(float(deviation_value)))
            if comment_id is None:
                insert_comment(payload2)
            else:
                update_comment(comment_id, payload2)
            return
        raise


# -----------------------------
# Excel Generation Logic
# -----------------------------

def generate_excel_template(
    data: pd.DataFrame, 
    start_date: date, 
    end_date: date
) -> bytes:
    """
    Generates an Excel file in memory with:
    - Pre-filled columns: site_name, deviation, equipment_name, start_date, end_date
    - User-fillable: reasons, remarks
    - Data Validation: Dropdown for 'reasons'
    """
    output = io.BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulk Upload"

    # Define headers
    headers = ["site_name", "deviation", "equipment_name", "reasons", "start_date", "end_date", "remarks"]
    ws.append(headers)

    # Fill data
    for _, row in data.iterrows():
        ws.append([
            row["site_name"],
            row["deviation"],
            row["equipment_name"],
            "",  # User fills reason
            start_date,
            end_date,
            ""   # User fills remarks
        ])

    # Create Data Validation for 'reasons' column (Col D = 4)
    # Excel validation lists must be comma-separated string if simple, or reference a range.
    # Since REASONS list is long, we'll put it in a separate hidden sheet and reference it.
    
    validation_sheet = wb.create_sheet("RefData")
    validation_sheet.sheet_state = "hidden"
    for i, r in enumerate(REASONS, 1):
        validation_sheet.cell(row=i, column=1, value=r)
    
    # Define range for validation: RefData!$A$1:$A$N
    ref_formula = f"RefData!$A$1:$A${len(REASONS)}"
    
    dv = DataValidation(type="list", formula1=ref_formula, allow_blank=True)
    dv.error = "Please select a valid reason from the list"
    dv.errorTitle = "Invalid Reason"
    dv.prompt = "Select a reason"
    dv.promptTitle = "Reason Selection"

    # Apply validation to 'reasons' column (D), rows 2 to 5000
    ws.add_data_validation(dv)
    dv.add(f"D2:D{max(len(data) + 100, 500)}")

    # Adjust column widths
    for col_idx, col_name in enumerate(headers, 1):
        column_letter = get_column_letter(col_idx)
        ws.column_dimensions[column_letter].width = 20

    wb.save(output)
    return output.getvalue()


# -----------------------------
# UI & Logic
# -----------------------------


def _format_equipment_label(equipment_name: str, dev: float) -> str:
    return f"{dev:.2f}%_{equipment_name}"


def _reset_comment_form_state(*, default_site: Optional[str], dmin: date, dmax: date) -> None:
    st.session_state["comments_edit_id"] = None
    st.session_state["comments_edit_row"] = None
    st.session_state["ac_pending_reset"] = {
        "default_site": default_site,
        "dmin": dmin,
        "dmax": dmax,
    }


def _ensure_comment_form_state(*, dmin: date, dmax: date) -> None:
    st.session_state.setdefault("ac_site", None)
    st.session_state.setdefault("ac_threshold", -3.0)
    st.session_state.setdefault("ac_from", None)
    st.session_state.setdefault("ac_to", None)
    st.session_state.setdefault("ac_equipment_labels", [])
    st.session_state.setdefault("ac_reasons", [])
    st.session_state.setdefault("ac_remarks", "")
    st.session_state.setdefault("ac_pending_reset", None)
    
    # Upload state
    st.session_state.setdefault("up_site", None)
    st.session_state.setdefault("up_thresh", -3.0)
    st.session_state.setdefault("up_from", None)
    st.session_state.setdefault("up_to", None)
    st.session_state.setdefault("up_pending_reset", None)
    st.session_state.setdefault("up_file_counter", 0)  # Counter for file uploader key


def _render_aggrid_table(df: pd.DataFrame, *, key: str, height: int = 380) -> None:
    if df is None or df.empty:
        st.info("No comments found.")
        return

    if not _HAS_AGGRID:
        st.dataframe(df, use_container_width=True, hide_index=True, height=height)
        return

    gb = GridOptionsBuilder.from_dataframe(df)
    gb.configure_default_column(sortable=True, filter=True, resizable=True, wrapText=True, autoHeight=True)
    gb.configure_grid_options(domLayout="normal")
    grid_options = gb.build()

    custom_css = {
        ".ag-row-hover": {"background-color": "rgba(255, 179, 0, 0.14) !important"},
        ".ag-theme-balham": {"font-family": "Inter, ui-sans-serif, system-ui, -apple-system, Segoe UI, Roboto, Arial, sans-serif"},
        ".ag-header-cell-label": {"font-weight": "700"},
    }

    AgGrid(
        df,
        gridOptions=grid_options,
        theme="balham",
        height=height,
        fit_columns_on_grid_load=True,
        allow_unsafe_jscode=False,
        custom_css=custom_css,
        enable_enterprise_modules=False,
        update_mode=GridUpdateMode.NO_UPDATE,
        key=key,
    )


def render(db_path: str) -> None:
    st.markdown("## Add Comments")
    st.caption("Add deviation-based comments for underperforming equipment to guide operations.")

    sites = list_sites_from_syd(db_path)
    if not sites:
        st.error("No sites found in `syd`. Load data first.")
        return

    dmin, dmax = date_bounds_from_syd(db_path)
    if not dmin or not dmax:
        st.error("No dates found in `syd`.")
        return

    st.session_state.setdefault("comments_edit_id", None)
    st.session_state.setdefault("comments_edit_row", None)
    
    _ensure_comment_form_state(dmin=dmin, dmax=dmax)

    # Apply pending reset for manual entry
    pending = st.session_state.get("ac_pending_reset")
    if isinstance(pending, dict):
        st.session_state["ac_site"] = None
        st.session_state["ac_threshold"] = -3.0
        st.session_state["ac_from"] = None
        st.session_state["ac_to"] = None
        st.session_state["ac_equipment_labels"] = []
        st.session_state["ac_reasons"] = []
        st.session_state["ac_remarks"] = ""
        st.session_state["ac_pending_reset"] = None
        st.session_state["_ac_prefilled"] = False
        st.session_state["_ac_equipment_prefilled"] = False

    # Apply pending reset for bulk upload (BEFORE widgets are created)
    up_pending = st.session_state.get("up_pending_reset")
    if up_pending is True:
        st.session_state["up_site"] = None
        st.session_state["up_thresh"] = -3.0
        st.session_state["up_from"] = None
        st.session_state["up_to"] = None
        # Increment file uploader counter to force new widget instance (file uploaders can't be reset via session state)
        st.session_state["up_file_counter"] = st.session_state.get("up_file_counter", 0) + 1
        st.session_state["up_pending_reset"] = None

    edit_row = st.session_state.get("comments_edit_row") or {}
    is_edit_mode = bool(st.session_state.get("comments_edit_id"))

    # ==========================
    # 1. MANUAL ENTRY EXPANDER
    # ==========================
    # Auto-expand if in edit mode, otherwise collapsed by default
    manual_expanded = is_edit_mode 
    with st.expander("Manual Entry", expanded=manual_expanded):
        
        # Prefill if edit mode
        if is_edit_mode and edit_row and not st.session_state.get("_ac_prefilled", False):
            try:
                st.session_state["ac_site"] = str(edit_row.get("site_name") or sites[0])
                st.session_state["ac_threshold"] = float(edit_row.get("threshold") or 2.0)
                st.session_state["ac_from"] = date.fromisoformat(str(edit_row.get("start_date"))) if edit_row.get("start_date") else dmin
                st.session_state["ac_to"] = date.fromisoformat(str(edit_row.get("end_date"))) if edit_row.get("end_date") else dmax
                st.session_state["ac_reasons"] = list(edit_row.get("reasons") or [])
                st.session_state["ac_remarks"] = str(edit_row.get("remarks") or "")
            except Exception:
                pass
            st.session_state["_ac_prefilled"] = True
        
        if is_edit_mode:
            st.info("Edit mode: update fields and click **Update**.")

        c1, c2, c3 = st.columns([2.4, 1.2, 2.4])
        with c1:
            site_name = st.selectbox(
                "Site Name", 
                options=["(select)", *sites], 
                index=0 if st.session_state.get("ac_site") is None else (sites.index(st.session_state["ac_site"]) + 1 if st.session_state["ac_site"] in sites else 0), 
                key="ac_site"
            )
            if site_name == "(select)": site_name = None
        with c2:
            threshold = st.number_input("Threshold (%)", min_value=-100.0, max_value=100.0, step=0.1, key="ac_threshold")
        with c3:
            dcol1, dcol2 = st.columns(2)
            with dcol1: start_date = st.date_input("From", min_value=dmin, max_value=dmax, key="ac_from")
            with dcol2: end_date = st.date_input("To", min_value=dmin, max_value=dmax, key="ac_to")

        if start_date and end_date and start_date > end_date:
            st.error("From date cannot be after To date.")

        # Candidate logic
        cand = pd.DataFrame()
        if site_name and start_date and end_date and threshold is not None:
            try:
                cand = equipment_deviation_candidates(db_path, site_name=str(site_name), start_date=start_date, end_date=end_date, threshold=float(threshold))
            except Exception as e:
                st.error(f"Failed to fetch equipment: {e}")

        label_to_equipment = {}
        label_to_dev = {}
        if not cand.empty:
            for _, r in cand.iterrows():
                eq = str(r["equipment_name"])
                dev = float(r["syd_dev_pct"])
                lbl = _format_equipment_label(eq, dev)
                label_to_equipment[lbl] = eq
                label_to_dev[lbl] = dev
        
        # Prefill equipment for edit
        if is_edit_mode and edit_row and not st.session_state.get("_ac_equipment_prefilled", False):
            try:
                desired = [str(x) for x in (edit_row.get("equipment_names") or [])]
                selected_labels = []
                for eq in desired:
                    matches = [l for l in label_to_equipment.keys() if l.endswith(f"_{eq}") or l.endswith(f"%_{eq}")]
                    if matches: selected_labels.append(matches[0])
                st.session_state["ac_equipment_labels"] = selected_labels
            except Exception: pass
            st.session_state["_ac_equipment_prefilled"] = True

        equipment_labels = st.multiselect("Equipment Name", options=list(label_to_equipment.keys()), key="ac_equipment_labels")
        reasons = st.multiselect("Reason(s)", options=REASONS, key="ac_reasons")
        remarks = st.text_area("Remarks", key="ac_remarks", height=100)

        if equipment_labels:
            devs = [float(label_to_dev.get(lbl, 0.0)) for lbl in equipment_labels]
            avg_dev = float(sum(devs) / max(len(devs), 1))
            st.caption(f"Selected: {len(equipment_labels)} | Avg Deviation: {avg_dev:.2f}%")

        if st.button("Update" if is_edit_mode else "Add Comment", type="primary"):
            if not site_name or not start_date or not end_date:
                st.error("Missing required fields.")
            elif not equipment_labels:
                st.error("Select equipment.")
            else:
                eq_names = [label_to_equipment[lbl] for lbl in equipment_labels]
                deviations = [float(label_to_dev.get(lbl, 0.0)) for lbl in equipment_labels]
                deviation_val = float(sum(deviations) / max(len(deviations), 1))
                
                payload = {
                    "site_name": str(site_name),
                    "start_date": _to_iso(start_date),
                    "end_date": _to_iso(end_date),
                    "equipment_names": eq_names,
                    "reasons": reasons,
                    "remarks": remarks,
                    "deviation": round(deviation_val, 6)
                }
                try:
                    comment_id = st.session_state.get("comments_edit_id")
                    _write_comment_with_fallback(comment_id=comment_id, payload=payload, deviation_value=deviation_val)
                    st.success("Success!")
                    _reset_comment_form_state(default_site=None, dmin=dmin, dmax=dmax)
                    st.rerun()
                except Exception as e:
                    st.error(f"Error: {e}")

    # ==========================
    # 2. BULK UPLOAD EXPANDER
    # ==========================
    with st.expander("Bulk Upload", expanded=False):
        st.info("Download data for underperforming equipment, fill in 'reasons' and 'remarks', then upload.")
        
        # --- Download Section ---
        c_up1, c_up2, c_up3 = st.columns([2, 1, 2])
        with c_up1:
            up_site = st.selectbox("Site Name", options=["(select)", *sites], key="up_site")
        with c_up2:
            up_thresh = st.number_input("Threshold (%)", value=-3.0, step=0.1, key="up_thresh")
        with c_up3:
            uc1, uc2 = st.columns(2)
            # Avoid widget warning: only set default value if session state doesn't have it
            up_from_val = st.session_state.get("up_from") if "up_from" in st.session_state else dmin
            up_to_val = st.session_state.get("up_to") if "up_to" in st.session_state else dmax
            with uc1: up_from = st.date_input("From", value=up_from_val, min_value=dmin, max_value=dmax, key="up_from")
            with uc2: up_to = st.date_input("To", value=up_to_val, min_value=dmin, max_value=dmax, key="up_to")

        download_df = pd.DataFrame()
        if up_site != "(select)" and up_from and up_to:
            # Check button click
            try:
                download_df = fetch_download_candidates(
                    db_path, 
                    site_name=str(up_site), 
                    start_date=up_from, 
                    end_date=up_to, 
                    threshold=float(up_thresh)
                )
            except Exception as e:
                st.error(f"Error fetching data: {e}")

        if not download_df.empty:
            st.write(f"Found {len(download_df)} equipment(s) below {up_thresh}%.")
            excel_bytes = generate_excel_template(download_df, up_from, up_to)
            
            st.download_button(
                label="Download Excel Template",
                data=excel_bytes,
                file_name=f"Upload_{up_site}_{up_from}_{up_to}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="btn_download_template"
            )
        elif up_site != "(select)":
            st.caption("No equipment found matching criteria.")

        st.markdown("---")
        
        # --- Upload Section ---
        # Use dynamic key based on counter to allow reset (file uploaders can't be reset via session state)
        file_uploader_key = f"up_file_{st.session_state.get('up_file_counter', 0)}"
        uploaded_file = st.file_uploader("Upload Filled Excel", type=["xlsx"], key=file_uploader_key)
        
        if uploaded_file:
            try:
                # Use engine='openpyxl' specifically to handle data validation warnings gracefully
                df_up = pd.read_excel(uploaded_file, engine="openpyxl")
                
                # Basic validation
                req_cols = ["site_name", "deviation", "equipment_name", "reasons", "start_date", "end_date", "remarks"]
                missing = [c for c in req_cols if c not in df_up.columns]
                
                if missing:
                    st.error(f"Invalid format. Missing columns: {missing}")
                else:
                    # Validate all rows first before showing submit button
                    validation_results = []
                    valid_payloads = []
                    invalid_rows = []
                    
                    for i, row in df_up.iterrows():
                        row_num = i + 2  # Excel row number (accounting for header)
                        errors = []
                        
                        # Extract values
                        s_name = str(row["site_name"]) if pd.notna(row["site_name"]) else ""
                        eq_name = str(row["equipment_name"]) if pd.notna(row["equipment_name"]) else ""
                        reason_raw = str(row["reasons"]) if pd.notna(row["reasons"]) else ""
                        remark_raw = str(row["remarks"]) if pd.notna(row["remarks"]) else ""
                        dev_val = row["deviation"] if pd.notna(row["deviation"]) else None
                        
                        # Validate required fields
                        if not reason_raw.strip():
                            errors.append("Reason is missing")
                        
                        if not remark_raw.strip():
                            errors.append("Remarks is missing")
                        
                        # Validate dates
                        sd = None
                        ed = None
                        try:
                            if pd.notna(row["start_date"]):
                                sd = pd.to_datetime(row["start_date"]).strftime("%Y-%m-%d")
                            else:
                                errors.append("Start date is missing")
                        except Exception:
                            errors.append("Start date is invalid")
                        
                        try:
                            if pd.notna(row["end_date"]):
                                ed = pd.to_datetime(row["end_date"]).strftime("%Y-%m-%d")
                            else:
                                errors.append("End date is missing")
                        except Exception:
                            errors.append("End date is invalid")
                        
                        # Validate deviation
                        if dev_val is None:
                            errors.append("Deviation is missing")
                        else:
                            try:
                                dev_val = float(dev_val)
                            except (ValueError, TypeError):
                                errors.append("Deviation is invalid")
                                dev_val = None
                        
                        # If all valid, add to payloads
                        if not errors and sd and ed and dev_val is not None:
                            valid_payloads.append({
                                "site_name": s_name,
                                "start_date": sd,
                                "end_date": ed,
                                "equipment_names": [eq_name],
                                "reasons": [reason_raw],
                                "remarks": remark_raw,
                                "deviation": round(dev_val, 6)
                            })
                        else:
                            invalid_rows.append({
                                "row": row_num,
                                "equipment": eq_name,
                                "errors": errors
                            })
                    
                    # Show validation summary
                    total_rows = len(df_up)
                    valid_count = len(valid_payloads)
                    invalid_count = len(invalid_rows)
                    
                    if invalid_count == 0:
                        st.success(f"✅ All good! All {valid_count} row(s) are valid and ready to submit.")
                    else:
                        st.warning(f"⚠️ Validation: {valid_count} valid row(s), {invalid_count} invalid row(s) will be rejected.")
                        
                        # Show invalid rows details
                        with st.expander(f"View {invalid_count} rejected row(s)", expanded=False):
                            for inv in invalid_rows:
                                st.error(f"Row {inv['row']} (Equipment: {inv['equipment']}): {', '.join(inv['errors'])}")
                    
                    # Only show submit button if there are valid rows
                    if valid_count > 0:
                        if st.button("Submit to Supabase", type="primary"):
                            progress_bar = st.progress(0)
                            status_text = st.empty()
                            
                            try:
                                insert_bulk_comments(valid_payloads)
                                progress_bar.progress(100)
                                status_text.text("Upload Complete.")
                                
                                # Success message that persists
                                placeholder = st.empty()
                                placeholder.success(f"Successfully submitted {len(valid_payloads)} record(s)!")
                                time.sleep(10)  # Wait 10 seconds before resetting form
                                
                                # Set pending reset flag (will be applied on next rerun, before widgets are created)
                                st.session_state["up_pending_reset"] = True
                                
                                placeholder.empty()
                                st.rerun()
                                
                            except Exception as e:
                                st.error(f"Supabase submission failed: {e}")
                    else:
                        st.error("❌ No valid rows to submit. Please fix the errors and re-upload.")

            except Exception as e:
                st.error(f"Failed to process file: {e}")


    # ==========================
    # 3. EXISTING COMMENTS & EDIT UI
    # ==========================
    st.write("")
    with st.expander("Edit existing comment", expanded=False):
        try:
            existing_for_edit = fetch_comments_live(limit=500)
        except Exception:
            existing_for_edit = pd.DataFrame()

        if existing_for_edit.empty:
            st.info("No comments found.")
        else:
            def _label_row(r):
                sid = str(r.get("site_name",""))
                sd = str(r.get("start_date",""))
                eq_raw = r.get("equipment_names")
                if isinstance(eq_raw, list):
                    eq = ", ".join(str(v) for v in eq_raw)
                else:
                    eq = str(eq_raw) if eq_raw is not None else ""
                return f"{sid} | {sd} | {eq}"
            
            labels = [_label_row(existing_for_edit.iloc[i]) for i in range(len(existing_for_edit))]
            label_to_idx = {labels[i]: i for i in range(len(labels))}
            picked = st.selectbox("Pick comment", ["(none)", *labels], key="ac_edit_pick")

            c_act, c_note = st.columns([1.6, 8.4])
            with c_act:
                if st.button("Load", disabled=(picked=="(none)"), key="ac_edit_load"):
                    row = existing_for_edit.iloc[label_to_idx[picked]].to_dict()
                    st.session_state["comments_edit_id"] = row.get("id")
                    st.session_state["comments_edit_row"] = row
                    st.rerun()
                if st.button("Clear edit mode", disabled=not is_edit_mode, key="ac_edit_clear"):
                    _reset_comment_form_state(default_site=None, dmin=dmin, dmax=dmax)
                    st.rerun()

    st.write("")
    with st.expander("Existing Comments (table)", expanded=False):
        f1, f2, f3 = st.columns([2.2, 2.0, 2.0])
        with f1: f_site = st.selectbox("Filter: Site", ["(all)", *sites], key="ac_exist_site")
        with f2: f_from = st.date_input("Filter: From", value=None, min_value=dmin, max_value=dmax, key="ac_exist_from")
        with f3: f_to = st.date_input("Filter: To", value=None, min_value=dmin, max_value=dmax, key="ac_exist_to")
        
        fs = None if f_site == "(all)" else f_site
        try:
            base = fetch_comments_live(site_name=fs, start_date=f_from, end_date=f_to, limit=500)
            if not base.empty:
                # Format equipment_names and reasons columns (convert lists to comma-separated strings)
                if "equipment_names" in base.columns:
                    base["equipment_names"] = base["equipment_names"].apply(
                        lambda x: ", ".join(str(v) for v in x) if isinstance(x, list) else (str(x) if x is not None else "")
                    )
                if "reasons" in base.columns:
                    base["reasons"] = base["reasons"].apply(
                        lambda x: ", ".join(str(v) for v in x) if isinstance(x, list) else (str(x) if x is not None else "")
                    )
                # Remove columns that should not be displayed
                columns_to_hide = ["id", "created_by", "created_at"]
                display_df = base.drop(columns=[c for c in columns_to_hide if c in base.columns])
                _render_aggrid_table(display_df, key="ac_existing_comments")
            else:
                st.info("No records.")
        except Exception as e:
            st.error(str(e))
