"""
store_data_table_duckdb.py

Scans DGR/ for Excel files and loads cleaned data into a single DuckDB ("master").

Rules implemented (per requirements):
- site_name extracted from filename (strip leading "DGR_" and extension)
- missing/blank/non-numeric/erroneous numeric values -> stored as 0
- Duplicate detection is FULL-ROW based (all columns must match exactly).
  - If exact duplicate rows are detected (all columns match an existing row),
    the script will pause up to 60 seconds waiting for terminal input:
      type "replace all" to delete+reinsert those exact duplicates
      otherwise exact duplicates are skipped and only non-duplicates are loaded

Note on updates:
- Tables have primary keys for integrity. Rows that share a PK but differ in any
  metric are NOT considered duplicates and will be updated via MERGE.
"""

from __future__ import annotations

import argparse
import logging
import os
import sys
import time
from dataclasses import dataclass
from datetime import date
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import duckdb
import pandas as pd
from openpyxl import load_workbook

try:
    from tqdm import tqdm
except Exception:  # pragma: no cover
    tqdm = None


LOGGER = logging.getLogger("dgr_loader")


# -----------------------------
# Column mappings (Excel -> DB)
# -----------------------------

DAILY_KPI_SHEET = "Daily_KPI"
SYD_SHEET = "Inv_SY_D"
PR_SHEET = "Inv_PR"


@dataclass(frozen=True)
class DailyKpiMapping:
    # db_col -> excel_col
    mapping: Dict[str, str]


@dataclass(frozen=True)
class BudgetKpiMapping:
    mapping: Dict[str, str]


DAILY_KPI_MAPPING = DailyKpiMapping(
    mapping={
        "date": "Date",
        "days": "Days",
        "ghi": "GHI-UP (KWh/m2)",
        "poa": "POA-UP(KWh/m2)",
        "at": "Amb_Temp(°C)",
        "mt": "Mod_Temp(°C)",
        "wsa": "WS_Avg(m/s)",
        "wsm": "WS_Max(m/s)",
        "pa_percent": "E_PA(%)",
        "ga_percent": "E_EGA(%)",
        "pr_percent": "PR(%)",
        "ac_cuf_percent": "AC_CUF(%)",
        "dc_cuf_percent": "DC_CUF(%)",
        "inv_gen_kwh": "Gen_Exp (kWh)",
        "abt_export_kwh": "Mtr_Export (kWh)",
        "abt_import_kwh": "Mtr_Import (kWh)",
        "abt_net_export_kwh": "Mtr_Net_Exp (KWh)",
        "operational_ac_mw": "Operational AC Capacity (MW)",
        "operational_dc_mwp": "Operational DC Capacity (MWp)",
    }
)

BUDGET_KPI_MAPPING = BudgetKpiMapping(
    mapping={
        "date": "Date",
        "b_poa": "Bugt_Resource",
        "b_energy_kwh": "Bugt_Energy",
        "b_pr_percent": "Bugt PR",
        "b_cuf_percent": "Bugt CUF",
        "b_pa_percent": "Bugt_PA",
        "b_ga_percent": "Bugt_EGA",
        "b_ac_mw": "Bugt AC Capacity (MW)",
        "b_dc_mwp": "Bugt Capacity",
        "b_sl_percent": "Budg Soiling Loss (%)",
    }
)


# -----------------------------
# Helpers
# -----------------------------


def _site_name_from_path(xlsx_path: Path) -> str:
    stem = xlsx_path.stem
    if stem.upper().startswith("DGR_"):
        return stem[4:]
    LOGGER.warning("Filename does not start with DGR_: %s (using %s)", xlsx_path.name, stem)
    return stem


def _safe_number(x: Any) -> float:
    """
    Convert blanks/non-numerics/errors to 0, otherwise float.
    """
    if x is None:
        return 0.0
    if isinstance(x, (int, float)) and pd.notna(x):
        # Convert NaN to 0
        try:
            if pd.isna(x):
                return 0.0
        except Exception:
            pass
        return float(x)
    if isinstance(x, str):
        s = x.strip()
        if s == "" or s.upper() in {"NA", "N/A", "NULL", "NONE", "-"}:
            return 0.0
        # Remove commas in numbers like "1,234.5"
        s = s.replace(",", "")
        try:
            return float(s)
        except Exception:
            return 0.0
    try:
        if pd.isna(x):
            return 0.0
    except Exception:
        pass
    try:
        return float(x)
    except Exception:
        return 0.0


def _to_date_series(s: pd.Series) -> pd.Series:
    # Robust date parsing from Excel datetimes/strings
    dt = pd.to_datetime(s, errors="coerce")
    return dt.dt.date


def _normalize_columns(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    return df


def _read_excel_sheet_safe(path: Path, sheet_name: str, *, header: int = 0) -> Optional[pd.DataFrame]:
    try:
        # Prefer calamine (Rust-based) for speed; fall back to openpyxl.
        try:
            df = pd.read_excel(path, sheet_name=sheet_name, engine="calamine", header=header)
        except Exception:
            df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl", header=header)
        return _normalize_columns(df)
    except ValueError as e:
        # Sheet missing
        LOGGER.warning("Missing sheet '%s' in %s: %s", sheet_name, path.name, e)
        return None
    except Exception as e:
        LOGGER.exception("Failed reading sheet '%s' in %s: %s", sheet_name, path.name, e)
        return None


def _find_date_column(columns: Sequence[str]) -> Optional[str]:
    for c in columns:
        if str(c).strip().lower() == "date":
            return c
    # fallback contains date
    for c in columns:
        if "date" in str(c).strip().lower():
            return c
    return None


def _coerce_numeric_columns(df: pd.DataFrame, cols: Iterable[str]) -> pd.DataFrame:
    df = df.copy()
    for c in cols:
        if c in df.columns:
            df[c] = df[c].map(_safe_number).astype("float64")
        else:
            df[c] = 0.0
    return df


# -----------------------------
# Extractors
# -----------------------------


def extract_daily_and_budget(path: Path, site_name: str) -> Tuple[pd.DataFrame, pd.DataFrame, List[date]]:
    df = _read_excel_sheet_safe(path, DAILY_KPI_SHEET, header=0)
    if df is None or df.empty:
        return (
            pd.DataFrame(columns=["site_name"] + list(DAILY_KPI_MAPPING.mapping.keys())),
            pd.DataFrame(columns=["site_name"] + ["b_ghi"] + list(BUDGET_KPI_MAPPING.mapping.keys())),
            [],
        )

    # Build daily df
    daily = pd.DataFrame()
    for db_col, excel_col in DAILY_KPI_MAPPING.mapping.items():
        if excel_col in df.columns:
            daily[db_col] = df[excel_col]
        else:
            LOGGER.warning("Missing column '%s' in %s/%s (daily_kpi) -> 0", excel_col, path.name, DAILY_KPI_SHEET)
            daily[db_col] = None
    daily["site_name"] = site_name

    # Parse dates, drop rows without a date (can't key)
    daily["date"] = _to_date_series(daily["date"])
    before = len(daily)
    daily = daily[daily["date"].notna()].copy()
    if len(daily) != before:
        LOGGER.warning("Dropped %d rows with invalid/missing dates in %s/%s", before - len(daily), path.name, DAILY_KPI_SHEET)

    numeric_cols = [c for c in daily.columns if c not in {"site_name", "date"}]
    daily = _coerce_numeric_columns(daily, numeric_cols)

    # Budget df (same sheet)
    budget = pd.DataFrame()
    for db_col, excel_col in BUDGET_KPI_MAPPING.mapping.items():
        if excel_col in df.columns:
            budget[db_col] = df[excel_col]
        else:
            LOGGER.warning("Missing column '%s' in %s/%s (budget_kpi) -> 0", excel_col, path.name, DAILY_KPI_SHEET)
            budget[db_col] = None
    budget["b_ghi"] = 0.0  # placeholder per requirement
    budget["site_name"] = site_name

    budget["date"] = _to_date_series(budget["date"])
    before = len(budget)
    budget = budget[budget["date"].notna()].copy()
    if len(budget) != before:
        LOGGER.warning("Dropped %d budget rows with invalid/missing dates in %s/%s", before - len(budget), path.name, DAILY_KPI_SHEET)

    budget_numeric = [c for c in budget.columns if c not in {"site_name", "date"}]
    budget = _coerce_numeric_columns(budget, budget_numeric)

    dates = sorted({d for d in daily["date"].tolist() if isinstance(d, date)})
    return daily, budget, dates


def extract_syd(path: Path, site_name: str) -> pd.DataFrame:
    """
    Inv_SY_D: equipment names in row 3 starting at column C (C3, D3, ...)
    We read with header=2 (0-based row index) so row 3 becomes header row.
    """
    df = _read_excel_sheet_safe(path, SYD_SHEET, header=2)
    if df is None or df.empty:
        return pd.DataFrame(columns=["site_name", "date", "equipment_name", "syd_percent"])

    date_col = _find_date_column(df.columns)
    if not date_col:
        LOGGER.warning("Could not find Date column in %s/%s; skipping SYD", path.name, SYD_SHEET)
        return pd.DataFrame(columns=["site_name", "date", "equipment_name", "syd_percent"])

    df = df.copy()
    df[date_col] = _to_date_series(df[date_col])
    df = df[df[date_col].notna()].copy()

    # equipment columns: from column C onward (index 2)
    equip_cols = list(df.columns[2:]) if len(df.columns) > 2 else []
    if not equip_cols:
        LOGGER.warning("No equipment columns detected in %s/%s; skipping SYD", path.name, SYD_SHEET)
        return pd.DataFrame(columns=["site_name", "date", "equipment_name", "syd_percent"])

    long_df = df[[date_col] + equip_cols].melt(id_vars=[date_col], var_name="equipment_name", value_name="syd_percent")
    long_df.rename(columns={date_col: "date"}, inplace=True)
    long_df["site_name"] = site_name
    long_df["syd_percent"] = long_df["syd_percent"].map(_safe_number).astype("float64")
    long_df["equipment_name"] = long_df["equipment_name"].astype(str).str.strip()
    long_df = long_df[long_df["equipment_name"] != ""].copy()
    return long_df[["site_name", "date", "equipment_name", "syd_percent"]]


def extract_pr(path: Path, site_name: str) -> Tuple[pd.DataFrame, List[str]]:
    """
    Inv_PR: equipment names start from D3 onwards (D3, E3, ...)
    We read with header=2 so row 3 is header row; equipment columns start from index 3.
    """
    df = _read_excel_sheet_safe(path, PR_SHEET, header=2)
    if df is None or df.empty:
        return pd.DataFrame(columns=["site_name", "date", "equipment_name", "pr_percent"]), []

    date_col = _find_date_column(df.columns)
    if not date_col:
        LOGGER.warning("Could not find Date column in %s/%s; skipping PR", path.name, PR_SHEET)
        return pd.DataFrame(columns=["site_name", "date", "equipment_name", "pr_percent"]), []

    df = df.copy()
    df[date_col] = _to_date_series(df[date_col])
    df = df[df[date_col].notna()].copy()

    equip_cols = list(df.columns[3:]) if len(df.columns) > 3 else []
    if not equip_cols:
        LOGGER.warning("No equipment columns detected in %s/%s; skipping PR", path.name, PR_SHEET)
        return pd.DataFrame(columns=["site_name", "date", "equipment_name", "pr_percent"]), []

    long_df = df[[date_col] + equip_cols].melt(id_vars=[date_col], var_name="equipment_name", value_name="pr_percent")
    long_df.rename(columns={date_col: "date"}, inplace=True)
    long_df["site_name"] = site_name
    long_df["pr_percent"] = long_df["pr_percent"].map(_safe_number).astype("float64")
    long_df["equipment_name"] = long_df["equipment_name"].astype(str).str.strip()
    long_df = long_df[long_df["equipment_name"] != ""].copy()
    equipment_names = sorted(set(long_df["equipment_name"].tolist()))
    return long_df[["site_name", "date", "equipment_name", "pr_percent"]], equipment_names


def extract_dc_capacity(path: Path, site_name: str, dates: Sequence[date]) -> pd.DataFrame:
    """
    DC capacity fetched from Inv_PR sheet "D1 onwards as per equipment_name".
    Implementation: take equipment names from row 3 (D3+). For each equipment column, scan rows 1..2
    (or until row 2) for the first numeric value; if none found, 0.
    Store capacity per (site_name, date, equipment_name).
    """
    if not dates:
        return pd.DataFrame(columns=["site_name", "date", "equipment_name", "dc_capacity_kwp"])

    try:
        wb = load_workbook(filename=path, data_only=True, read_only=True)
    except Exception as e:
        LOGGER.exception("Failed opening workbook %s for dc_capacity: %s", path.name, e)
        return pd.DataFrame(columns=["site_name", "date", "equipment_name", "dc_capacity_kwp"])

    if PR_SHEET not in wb.sheetnames:
        LOGGER.warning("Missing sheet '%s' in %s (dc_capacity) -> empty", PR_SHEET, path.name)
        return pd.DataFrame(columns=["site_name", "date", "equipment_name", "dc_capacity_kwp"])

    ws = wb[PR_SHEET]

    # equipment header row = 3; start col = 4 (D)
    equip_row = 3
    start_col = 4
    capacities: Dict[str, float] = {}

    col = start_col
    while True:
        equip_name = ws.cell(row=equip_row, column=col).value
        if equip_name is None or str(equip_name).strip() == "":
            # stop at first blank header cell (typical layout)
            break

        equip_name = str(equip_name).strip()

        cap_val = 0.0
        for r in (1, 2):
            v = ws.cell(row=r, column=col).value
            num = _safe_number(v)
            if num != 0.0:
                cap_val = num
                break
        capacities[equip_name] = cap_val
        col += 1

    if not capacities:
        LOGGER.warning("No equipment headers found at row 3 from col D in %s/%s", path.name, PR_SHEET)
        return pd.DataFrame(columns=["site_name", "date", "equipment_name", "dc_capacity_kwp"])

    # Build df by replicating capacities for each date
    rows: List[Dict[str, Any]] = []
    for d in dates:
        for equip_name, cap in capacities.items():
            rows.append(
                {
                    "site_name": site_name,
                    "date": d,
                    "equipment_name": equip_name,
                    "dc_capacity_kwp": float(cap),
                }
            )
    return pd.DataFrame(rows, columns=["site_name", "date", "equipment_name", "dc_capacity_kwp"])


# -----------------------------
# DuckDB load/upsert
# -----------------------------


def ensure_schema(con: duckdb.DuckDBPyConnection) -> None:
    def q(ident: str) -> str:
        # Quote identifiers to avoid reserved-keyword issues (e.g., "at")
        return '"' + ident.replace('"', '""') + '"'

    con.execute(
        """
        CREATE TABLE IF NOT EXISTS daily_kpi (
          "site_name" TEXT NOT NULL,
          "date" DATE NOT NULL,
          "days" DOUBLE,
          "ghi" DOUBLE,
          "poa" DOUBLE,
          "at" DOUBLE,
          "mt" DOUBLE,
          "wsa" DOUBLE,
          "wsm" DOUBLE,
          "pa_percent" DOUBLE,
          "ga_percent" DOUBLE,
          "pr_percent" DOUBLE,
          "ac_cuf_percent" DOUBLE,
          "dc_cuf_percent" DOUBLE,
          "inv_gen_kwh" DOUBLE,
          "abt_export_kwh" DOUBLE,
          "abt_import_kwh" DOUBLE,
          "abt_net_export_kwh" DOUBLE,
          "operational_ac_mw" DOUBLE,
          "operational_dc_mwp" DOUBLE,
          CONSTRAINT daily_kpi_pk PRIMARY KEY("site_name", "date")
        );
        """
    )

    con.execute(
        """
        CREATE TABLE IF NOT EXISTS budget_kpi (
          "site_name" TEXT NOT NULL,
          "date" DATE NOT NULL,
          "b_poa" DOUBLE,
          "b_ghi" DOUBLE,
          "b_energy_kwh" DOUBLE,
          "b_pr_percent" DOUBLE,
          "b_cuf_percent" DOUBLE,
          "b_pa_percent" DOUBLE,
          "b_ga_percent" DOUBLE,
          "b_ac_mw" DOUBLE,
          "b_dc_mwp" DOUBLE,
          "b_sl_percent" DOUBLE,
          CONSTRAINT budget_kpi_pk PRIMARY KEY("site_name", "date")
        );
        """
    )

    con.execute(
        """
        CREATE TABLE IF NOT EXISTS syd (
          "site_name" TEXT NOT NULL,
          "date" DATE NOT NULL,
          "equipment_name" TEXT NOT NULL,
          "syd_percent" DOUBLE,
          CONSTRAINT syd_pk PRIMARY KEY("site_name", "date", "equipment_name")
        );
        """
    )

    con.execute(
        """
        CREATE TABLE IF NOT EXISTS pr (
          "site_name" TEXT NOT NULL,
          "date" DATE NOT NULL,
          "equipment_name" TEXT NOT NULL,
          "pr_percent" DOUBLE,
          CONSTRAINT pr_pk PRIMARY KEY("site_name", "date", "equipment_name")
        );
        """
    )

    con.execute(
        """
        CREATE TABLE IF NOT EXISTS dc_capacity (
          "site_name" TEXT NOT NULL,
          "date" DATE NOT NULL,
          "equipment_name" TEXT NOT NULL,
          "dc_capacity_kwp" DOUBLE,
          CONSTRAINT dc_capacity_pk PRIMARY KEY("site_name", "date", "equipment_name")
        );
        """
    )

    # Indexes (help lookups; PKs are already indexed-ish but DuckDB supports explicit indexes)
    con.execute(f"CREATE INDEX IF NOT EXISTS idx_daily_kpi_site_date ON daily_kpi({q('site_name')}, {q('date')});")
    con.execute(f"CREATE INDEX IF NOT EXISTS idx_budget_kpi_site_date ON budget_kpi({q('site_name')}, {q('date')});")
    con.execute(
        f"CREATE INDEX IF NOT EXISTS idx_syd_site_date_eq ON syd({q('site_name')}, {q('date')}, {q('equipment_name')});"
    )
    con.execute(
        f"CREATE INDEX IF NOT EXISTS idx_pr_site_date_eq ON pr({q('site_name')}, {q('date')}, {q('equipment_name')});"
    )
    con.execute(
        f"CREATE INDEX IF NOT EXISTS idx_dc_site_date_eq ON dc_capacity({q('site_name')}, {q('date')}, {q('equipment_name')});"
    )


def _merge_upsert(
    con: duckdb.DuckDBPyConnection,
    *,
    target_table: str,
    df: pd.DataFrame,
    key_cols: Sequence[str],
    value_cols: Sequence[str],
) -> int:
    if df is None or df.empty:
        return 0

    def q(ident: str) -> str:
        return '"' + ident.replace('"', '""') + '"'

    staging = f"stg_{target_table}"
    con.register(staging, df)

    on_clause = " AND ".join([f"t.{q(c)} = s.{q(c)}" for c in key_cols])
    set_clause = ", ".join([f"{q(c)} = s.{q(c)}" for c in value_cols])
    insert_cols = ", ".join([q(c) for c in list(key_cols) + list(value_cols)])
    insert_vals = ", ".join([f"s.{q(c)}" for c in list(key_cols) + list(value_cols)])

    distinct_predicates = [f"(t.{q(c)} IS DISTINCT FROM s.{q(c)})" for c in value_cols]
    distinct_where = " OR ".join(distinct_predicates) if distinct_predicates else "FALSE"

    sql = f"""
    MERGE INTO {target_table} AS t
    USING {staging} AS s
    ON {on_clause}
    WHEN MATCHED AND ({distinct_where})
      THEN UPDATE SET {set_clause}
    WHEN NOT MATCHED
      THEN INSERT ({insert_cols}) VALUES ({insert_vals});
    """

    con.execute(sql)
    # DuckDB doesn't return affected rowcount reliably for MERGE; return len(df) as processed
    return int(len(df))


def _timed_input(prompt: str, *, timeout_seconds: int = 60) -> Optional[str]:
    """
    Read a line from terminal with a timeout.

    Returns:
      - str: user input (without trailing newline)
      - None: timeout/no input

    Works on Windows (msvcrt) and falls back to a thread-based approach elsewhere.
    """
    # If stdin isn't interactive, don't hang; default to no input.
    try:
        if hasattr(sys.stdin, "isatty") and not sys.stdin.isatty():
            sys.stdout.write(prompt)
            sys.stdout.write("\n[stdin is not interactive] Defaulting to: add new (skip exact duplicates).\n")
            sys.stdout.flush()
            return None
    except Exception:
        pass

    try:
        import msvcrt  # type: ignore

        sys.stdout.write(prompt)
        sys.stdout.flush()
        start = time.time()
        buf: list[str] = []
        while (time.time() - start) < float(timeout_seconds):
            if msvcrt.kbhit():
                ch = msvcrt.getwch()
                if ch in ("\r", "\n"):
                    sys.stdout.write("\n")
                    sys.stdout.flush()
                    return "".join(buf).strip()
                if ch == "\b":
                    if buf:
                        buf.pop()
                        # erase char
                        sys.stdout.write("\b \b")
                        sys.stdout.flush()
                else:
                    buf.append(ch)
                    sys.stdout.write(ch)
                    sys.stdout.flush()
            time.sleep(0.05)
        sys.stdout.write("\n")
        sys.stdout.flush()
        return None
    except Exception:
        # Fallback: thread with join timeout
        import threading

        result: dict[str, Optional[str]] = {"value": None}

        def _read() -> None:
            try:
                result["value"] = sys.stdin.readline().strip()
            except Exception:
                result["value"] = None

        sys.stdout.write(prompt)
        sys.stdout.flush()
        t = threading.Thread(target=_read, daemon=True)
        t.start()
        t.join(timeout_seconds)
        return result["value"] if t.is_alive() is False else None


def _quote_ident(ident: str) -> str:
    return '"' + ident.replace('"', '""') + '"'


def _split_exact_duplicates(
    con: duckdb.DuckDBPyConnection,
    *,
    target_table: str,
    df: pd.DataFrame,
) -> tuple[pd.DataFrame, pd.DataFrame]:
    """
    Split incoming df into (exact_duplicates, non_duplicates) using FULL-ROW equality.
    A row is an exact duplicate only if ALL column values match an existing row.
    """
    if df is None or df.empty:
        return (pd.DataFrame(columns=list(df.columns) if df is not None else []), pd.DataFrame(columns=list(df.columns) if df is not None else []))

    cols = [str(c) for c in df.columns]
    staging = f"stg_fullrow_{target_table}"
    con.register(staging, df)

    pred = " AND ".join([f"t.{_quote_ident(c)} IS NOT DISTINCT FROM s.{_quote_ident(c)}" for c in cols]) or "TRUE"

    dup_sql = f"""
    SELECT s.*
    FROM {staging} s
    WHERE EXISTS (
      SELECT 1 FROM {target_table} t
      WHERE {pred}
    );
    """
    nondup_sql = f"""
    SELECT s.*
    FROM {staging} s
    WHERE NOT EXISTS (
      SELECT 1 FROM {target_table} t
      WHERE {pred}
    );
    """
    dups = con.execute(dup_sql).fetchdf()
    nondups = con.execute(nondup_sql).fetchdf()
    return (dups, nondups)


def _delete_exact_rows(
    con: duckdb.DuckDBPyConnection,
    *,
    target_table: str,
    df_exact: pd.DataFrame,
) -> None:
    """
    Delete rows from target_table that match df_exact by FULL-ROW equality.
    """
    if df_exact is None or df_exact.empty:
        return
    cols = [str(c) for c in df_exact.columns]
    staging = f"stg_delete_exact_{target_table}"
    con.register(staging, df_exact)
    pred = " AND ".join([f"t.{_quote_ident(c)} IS NOT DISTINCT FROM s.{_quote_ident(c)}" for c in cols]) or "TRUE"
    sql = f"""
    DELETE FROM {target_table} t
    WHERE EXISTS (
      SELECT 1 FROM {staging} s
      WHERE {pred}
    );
    """
    con.execute(sql)


def load_to_duckdb(db_path: Path, daily: pd.DataFrame, budget: pd.DataFrame, syd: pd.DataFrame, pr: pd.DataFrame, dc: pd.DataFrame) -> None:
    con = duckdb.connect(str(db_path))
    try:
        ensure_schema(con)

        steps = [
            (
                "daily_kpi",
                dict(
                    target_table="daily_kpi",
                    df=daily,
                    key_cols=("site_name", "date"),
                    value_cols=[c for c in daily.columns if c not in {"site_name", "date"}] if daily is not None else [],
                ),
            ),
            (
                "budget_kpi",
                dict(
                    target_table="budget_kpi",
                    df=budget,
                    key_cols=("site_name", "date"),
                    value_cols=[c for c in budget.columns if c not in {"site_name", "date"}] if budget is not None else [],
                ),
            ),
            (
                "syd",
                dict(
                    target_table="syd",
                    df=syd,
                    key_cols=("site_name", "date", "equipment_name"),
                    value_cols=["syd_percent"],
                ),
            ),
            (
                "pr",
                dict(
                    target_table="pr",
                    df=pr,
                    key_cols=("site_name", "date", "equipment_name"),
                    value_cols=["pr_percent"],
                ),
            ),
            (
                "dc_capacity",
                dict(
                    target_table="dc_capacity",
                    df=dc,
                    key_cols=("site_name", "date", "equipment_name"),
                    value_cols=["dc_capacity_kwp"],
                ),
            ),
        ]

        # 1) Full-row exact-duplicate detection for each table (ALL columns must match)
        dup_info: list[tuple[str, pd.DataFrame, pd.DataFrame, dict[str, Any]]] = []
        dup_tables: list[str] = []
        for name, kwargs in steps:
            df = kwargs.get("df")
            if df is None or df.empty:
                dup_info.append((name, pd.DataFrame(), pd.DataFrame(), kwargs))
                continue
            dups, nondups = _split_exact_duplicates(con, target_table=kwargs["target_table"], df=df)
            dup_info.append((name, dups, nondups, kwargs))
            if dups is not None and not dups.empty:
                dup_tables.append(name)

        # Terminal summary: redundant (=exact duplicates) vs new/changed (=non-duplicates)
        sys.stdout.write("\n=== DuckDB load pre-check (full-row duplicates) ===\n")
        for name, dups, nondups, kwargs in dup_info:
            df = kwargs.get("df")
            total = int(len(df)) if df is not None else 0
            dup_n = int(len(dups)) if dups is not None else 0
            non_n = int(len(nondups)) if nondups is not None else 0
            sys.stdout.write(f"- {name}: incoming={total} | exact_duplicates={dup_n} | to_load(new/changed)={non_n}\n")
        sys.stdout.flush()

        # 2) If any exact duplicates exist, prompt once for "replace all"
        replace_all = False
        if dup_tables:
            msg = (
                f"\nExact duplicate rows detected in: {', '.join(dup_tables)}.\n"
                f'Type "replace all" within 60 seconds to overwrite exact duplicates.\n'
                f"Else (timeout/anything else), duplicates will be skipped and only new/changed rows will be loaded.\n"
                f"[Waiting {60} seconds] > "
            )
            ans = _timed_input(msg, timeout_seconds=60)
            if ans is None or ans.strip() == "":
                sys.stdout.write("No input received within 60 seconds. Defaulting to: add new (skip exact duplicates).\n")
                sys.stdout.flush()
                replace_all = False
            else:
                replace_all = bool(ans and ans.strip().lower() == "replace all")
                if not replace_all:
                    sys.stdout.write('Input was not "replace all". Defaulting to: add new (skip exact duplicates).\n')
                    sys.stdout.flush()
            if replace_all:
                LOGGER.warning('User confirmed "replace all": deleting and reinserting exact duplicates.')
            else:
                LOGGER.info("No confirmation received: exact duplicates will be skipped.")
        else:
            sys.stdout.write("\nNo exact duplicates detected. Proceeding to load all incoming rows.\n")
            sys.stdout.flush()

        if tqdm is None:
            LOGGER.info("tqdm not installed; running DuckDB upserts without a progress bar.")
            for name, dups, nondups, kwargs in dup_info:
                # Replace mode: delete exact duplicates first, then MERGE full df
                # Skip mode: MERGE only non-duplicates (exact duplicates are ignored)
                df_to_load = kwargs.get("df")
                if df_to_load is None or df_to_load.empty:
                    continue
                if not dups.empty and replace_all:
                    _delete_exact_rows(con, target_table=kwargs["target_table"], df_exact=dups)
                    _merge_upsert(con, **kwargs)
                    sys.stdout.write(f"[{name}] replace all: deleted {len(dups)} exact duplicates; loaded {len(df_to_load)} rows.\n")
                else:
                    if nondups is None or nondups.empty:
                        if dups is not None and not dups.empty:
                            sys.stdout.write(f"[{name}] add new: skipped {len(dups)} exact duplicates; nothing new to load.\n")
                        continue
                    _merge_upsert(con, **{**kwargs, "df": nondups})
                    if dups is not None and not dups.empty:
                        sys.stdout.write(f"[{name}] add new: skipped {len(dups)} exact duplicates; loaded {len(nondups)} new/changed rows.\n")
                    else:
                        sys.stdout.write(f"[{name}] loaded {len(nondups)} rows.\n")
                sys.stdout.flush()
        else:
            with tqdm(total=len(steps), desc="Upserting into DuckDB", unit="table") as pbar:
                for name, dups, nondups, kwargs in dup_info:
                    LOGGER.info("Loading table %s ...", name)
                    df_to_load = kwargs.get("df")
                    if df_to_load is None or df_to_load.empty:
                        pbar.update(1)
                        continue
                    if not dups.empty and replace_all:
                        _delete_exact_rows(con, target_table=kwargs["target_table"], df_exact=dups)
                        _merge_upsert(con, **kwargs)
                        sys.stdout.write(f"[{name}] replace all: deleted {len(dups)} exact duplicates; loaded {len(df_to_load)} rows.\n")
                    else:
                        if nondups is None or nondups.empty:
                            if dups is not None and not dups.empty:
                                sys.stdout.write(f"[{name}] add new: skipped {len(dups)} exact duplicates; nothing new to load.\n")
                            pbar.update(1)
                            continue
                        _merge_upsert(con, **{**kwargs, "df": nondups})
                        if dups is not None and not dups.empty:
                            sys.stdout.write(f"[{name}] add new: skipped {len(dups)} exact duplicates; loaded {len(nondups)} new/changed rows.\n")
                        else:
                            sys.stdout.write(f"[{name}] loaded {len(nondups)} rows.\n")
                    sys.stdout.flush()
                    pbar.update(1)
    finally:
        con.close()


# -----------------------------
# Main
# -----------------------------


def configure_logging(verbose: bool) -> None:
    level = logging.DEBUG if verbose else logging.INFO
    logging.basicConfig(
        level=level,
        format="%(asctime)s | %(levelname)s | %(name)s | %(message)s",
    )


def iter_excel_files(dgr_dir: Path) -> List[Path]:
    if not dgr_dir.exists():
        raise FileNotFoundError(f"DGR directory not found: {dgr_dir}")
    files = sorted([p for p in dgr_dir.glob("*.xlsx") if p.is_file() and not p.name.startswith("~$")])
    return files


def main(argv: Optional[Sequence[str]] = None) -> int:
    parser = argparse.ArgumentParser(description="Load DGR Excel files into a master DuckDB.")
    parser.add_argument("--dgr-dir", default="DGR", help="Folder containing DGR_*.xlsx files (default: DGR)")
    parser.add_argument("--db", default="master.duckdb", help="Output DuckDB file path (default: master.duckdb)")
    parser.add_argument("--verbose", action="store_true", help="Enable debug logging")
    parser.add_argument("--no-progress", action="store_true", help="Disable progress bars")
    args = parser.parse_args(argv)

    configure_logging(args.verbose)

    dgr_dir = Path(args.dgr_dir)
    db_path = Path(args.db)

    files = iter_excel_files(dgr_dir)
    if not files:
        LOGGER.warning("No .xlsx files found in %s", dgr_dir)
        return 0

    all_daily: List[pd.DataFrame] = []
    all_budget: List[pd.DataFrame] = []
    all_syd: List[pd.DataFrame] = []
    all_pr: List[pd.DataFrame] = []
    all_dc: List[pd.DataFrame] = []

    file_iter: Iterable[Path]
    if (not args.no_progress) and tqdm is not None:
        file_iter = tqdm(files, desc="Processing Excel files", unit="file")
    else:
        if (not args.no_progress) and tqdm is None:
            LOGGER.info("tqdm not installed; running without file progress bar.")
        file_iter = files

    for f in file_iter:
        site = _site_name_from_path(f)
        LOGGER.info("Processing %s (site_name=%s)", f.name, site)

        daily, budget, dates = extract_daily_and_budget(f, site)
        syd = extract_syd(f, site)
        pr, _equip = extract_pr(f, site)
        dc = extract_dc_capacity(f, site, dates)

        all_daily.append(daily)
        all_budget.append(budget)
        all_syd.append(syd)
        all_pr.append(pr)
        all_dc.append(dc)

    # Avoid pandas concat warnings by excluding empties
    daily_df = pd.concat([d for d in all_daily if d is not None and not d.empty], ignore_index=True) if all_daily else pd.DataFrame()
    budget_df = (
        pd.concat([d for d in all_budget if d is not None and not d.empty], ignore_index=True) if all_budget else pd.DataFrame()
    )
    syd_df = pd.concat([d for d in all_syd if d is not None and not d.empty], ignore_index=True) if all_syd else pd.DataFrame()
    pr_df = pd.concat([d for d in all_pr if d is not None and not d.empty], ignore_index=True) if all_pr else pd.DataFrame()
    dc_df = pd.concat([d for d in all_dc if d is not None and not d.empty], ignore_index=True) if all_dc else pd.DataFrame()

    LOGGER.info(
        "Loaded frames: daily_kpi=%d, budget_kpi=%d, syd=%d, pr=%d, dc_capacity=%d",
        len(daily_df),
        len(budget_df),
        len(syd_df),
        len(pr_df),
        len(dc_df),
    )

    # Ensure types are friendly for DuckDB
    for df in (daily_df, budget_df, syd_df, pr_df, dc_df):
        if df is not None and not df.empty and "date" in df.columns:
            # pandas might store python date objects; DuckDB can ingest them
            pass

    load_to_duckdb(db_path, daily_df, budget_df, syd_df, pr_df, dc_df)
    LOGGER.info("Done. DuckDB written to %s", db_path.resolve())
    return 0


if __name__ == "__main__":
    raise SystemExit(main())


