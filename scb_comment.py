from __future__ import annotations

import io
import time
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Optional

import pandas as pd
import streamlit as st
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

import add_comments
import scb_ot


@dataclass(frozen=True)
class ScbOtRun:
    site_name: str
    threshold: float
    from_date: date
    to_date: date
    dev: pd.DataFrame
    dev_plot: pd.DataFrame
    available_dates: set[date]


def _now_utc_iso() -> str:
    # Supabase accepts ISO strings; use UTC without timezone suffix to match existing code style.
    return datetime.utcnow().replace(microsecond=0).isoformat()


def _safe_date(v: object) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, date) and not isinstance(v, datetime):
        return v
    if isinstance(v, datetime):
        return v.date()
    s = str(v).strip()
    if not s:
        return None
    try:
        return date.fromisoformat(s)
    except Exception:
        try:
            return pd.to_datetime(s, errors="coerce").date()  # type: ignore[union-attr]
        except Exception:
            return None


def _run_scb_ot_once(
    *,
    db_path: str,
    site_name: str,
    from_date: date,
    to_date: date,
    threshold: float,
    progress: Optional["st.delta_generator.DeltaGenerator"] = None,
) -> ScbOtRun:
    """
    IMPORTANT: SCB OT is the ONLY source of truth.
    This function uses the same internal SCB OT computation pipeline and then applies the same threshold filter.
    """
    if progress:
        progress.progress(0.05, text="Resolving site table…")

    table = scb_ot.resolve_site_table_name(db_path, site_name)
    cols = scb_ot.get_table_columns(db_path, table)
    scb_cols = [c for c in cols if str(c).upper().startswith("SCB")]
    if not scb_cols:
        raise RuntimeError("No SCB columns found in the selected site table.")

    # Availability dates (same logic as SCB OT; informational + used for validation/UI gating)
    available_dates: set[date] = set()
    try:
        available_dates = scb_ot._available_dates_for_site_table(db_path, table, tuple(scb_cols))  # type: ignore[attr-defined]
    except Exception:
        available_dates = set()

    if progress:
        progress.progress(0.20, text="Fetching raw SCB data…")

    df_raw = scb_ot._fetch_raw_scb_data(  # type: ignore[attr-defined]
        db_path=db_path,
        table=table,
        from_date=from_date,
        to_date=to_date,
        scb_cols=scb_cols,
    )
    if df_raw is None or df_raw.empty:
        raise RuntimeError("No SCB data found for the selected site/date window.")

    if progress:
        progress.progress(0.45, text="Computing SCB OT results…")

    dev, _remarks_df, abort_reason = scb_ot._compute_scb_ot_peak_pipeline(  # type: ignore[attr-defined]
        df_raw=df_raw,
        site_name=site_name,
        table=table,
        from_date=from_date,
        to_date=to_date,
        scb_cols=list(scb_cols),
        db_path=db_path,
    )
    if abort_reason:
        raise RuntimeError(f"SCB OT computation aborted: {abort_reason}")
    if dev is None or dev.empty:
        raise RuntimeError("SCB OT produced no results (all SCBs eliminated/skipped).")

    # Match SCB OT chart ordering: IS -> INV -> SCB
    dev = dev.copy()
    dev["sort_key"] = dev["scb_label"].apply(scb_ot._parse_scb_label)  # type: ignore[attr-defined]
    dev = dev.sort_values(["sort_key"], ascending=True).reset_index(drop=True)

    # Apply the SAME threshold filter rule used in SCB OT:
    dev_plot = dev[dev["deviation_pct"] <= float(threshold)].copy()

    if progress:
        progress.progress(0.60, text="Preparing final results…")

    return ScbOtRun(
        site_name=site_name,
        threshold=float(threshold),
        from_date=from_date,
        to_date=to_date,
        dev=dev,
        dev_plot=dev_plot,
        available_dates=available_dates,
    )


def _build_excel_template_bytes(*, run: ScbOtRun) -> bytes:
    """
    Excel Columns (Exact Order):
      site_name, scb_label, deviation_pct, disconnected_strings, reasons, start_date, end_date, remarks
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "SCB Comment"

    headers = ["site_name", "scb_label", "deviation_pct", "disconnected_strings", "reasons", "start_date", "end_date", "remarks"]
    ws.append(headers)

    df = run.dev_plot.copy()
    if df.empty:
        # Still generate a valid template with headers only.
        pass
    else:
        df = df.copy()
        df["deviation_pct"] = pd.to_numeric(df["deviation_pct"], errors="coerce")
        df["disconnected_strings"] = pd.to_numeric(df["disconnected_strings"], errors="coerce").fillna(0).astype(int)

        for _, r in df.iterrows():
            ws.append(
                [
                    run.site_name,
                    str(r.get("scb_label", "")),
                    float(r.get("deviation_pct", 0.0)) if pd.notna(r.get("deviation_pct")) else None,
                    int(r.get("disconnected_strings", 0)) if pd.notna(r.get("disconnected_strings")) else 0,
                    "",  # reasons (user must select)
                    run.from_date,
                    run.to_date,
                    "",  # remarks optional
                ]
            )

    # Hidden reference sheet for reasons list (no free text via dropdown validation).
    ref = wb.create_sheet("RefData")
    ref.sheet_state = "hidden"
    for i, reason in enumerate(add_comments.REASONS, 1):
        ref.cell(row=i, column=1, value=reason)

    ref_formula = f"RefData!$A$1:$A${len(add_comments.REASONS)}"
    dv = DataValidation(type="list", formula1=ref_formula, allow_blank=True)
    dv.error = "Please select a valid reason from the list"
    dv.errorTitle = "Invalid Reason"
    dv.prompt = "Select a reason"
    dv.promptTitle = "Reason Selection"

    # Apply validation to reasons column (D), rows 2..N
    ws.add_data_validation(dv)
    n_rows = max(ws.max_row, 2)
    # reasons is column E after adding site_name (A)
    dv.add(f"E2:E{max(n_rows + 200, 500)}")

    # Protect/lock the site_name column (A) while keeping other user fields editable.
    # Excel protection is lightweight; it prevents accidental edits and satisfies UX requirement.
    max_row = max(ws.max_row, 2)
    max_col = len(headers)
    # Unlock everything by default in the data grid
    for r in range(1, max_row + 1):
        for c in range(1, max_col + 1):
            ws.cell(row=r, column=c).protection = Protection(locked=False)
    # Lock site_name column (A)
    for r in range(1, max_row + 1):
        ws.cell(row=r, column=1).protection = Protection(locked=True)
    ws.protection.sheet = True

    # Basic column sizing for readability.
    widths = {
        "A": 14,  # site_name
        "B": 22,  # scb_label
        "C": 14,  # deviation_pct
        "D": 20,  # disconnected_strings
        "E": 28,  # reasons
        "F": 14,  # start_date
        "G": 14,  # end_date
        "H": 40,  # remarks
    }
    for col_letter, w in widths.items():
        ws.column_dimensions[col_letter].width = w

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


def _read_uploaded_excel(file_bytes: bytes) -> pd.DataFrame:
    wb = load_workbook(filename=io.BytesIO(file_bytes), data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return pd.DataFrame()

    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    data_rows = rows[1:]
    df = pd.DataFrame(data_rows, columns=header)

    # Drop fully empty rows
    df = df.dropna(how="all")
    return df


def _validate_upload(
    *,
    df_up: pd.DataFrame,
    run: ScbOtRun,
) -> tuple[list[dict[str, Any]], list[str]]:
    """
    Validates the uploaded Excel strictly and returns (payloads, errors).
    """
    errors: list[str] = []
    payloads: list[dict[str, Any]] = []

    if df_up is None or df_up.empty:
        return [], ["Uploaded Excel has no rows."]

    required_cols = ["site_name", "scb_label", "deviation_pct", "disconnected_strings", "reasons", "start_date", "end_date"]
    missing = [c for c in required_cols if c not in df_up.columns]
    if missing:
        return [], [f"Missing required columns: {missing}"]

    # System truth (from SCB OT run used for this submission)
    sys_df = run.dev_plot.copy()
    sys_df["deviation_pct"] = pd.to_numeric(sys_df["deviation_pct"], errors="coerce")
    sys_df["disconnected_strings"] = pd.to_numeric(sys_df["disconnected_strings"], errors="coerce").fillna(0).astype(int)

    sys_map_dev = {str(r["scb_label"]): float(r["deviation_pct"]) for _, r in sys_df.iterrows() if pd.notna(r.get("scb_label"))}
    sys_map_ds = {str(r["scb_label"]): int(r["disconnected_strings"]) for _, r in sys_df.iterrows() if pd.notna(r.get("scb_label"))}

    for idx, row in df_up.iterrows():
        excel_site = str(row.get("site_name") or "").strip()
        if not excel_site:
            errors.append(f"Row {idx+2}: site_name is blank.")
            continue
        if excel_site.strip().upper() != str(run.site_name).strip().upper():
            errors.append(f"Row {idx+2}: site_name '{excel_site}' does not match selected site '{run.site_name}'.")
            continue

        scb_label = str(row.get("scb_label") or "").strip()
        if not scb_label:
            errors.append(f"Row {idx+2}: scb_label is blank.")
            continue

        reason = str(row.get("reasons") or "").strip()
        if not reason:
            errors.append(f"Row {idx+2} ({scb_label}): reasons is mandatory.")
            continue
        if reason not in add_comments.REASONS:
            errors.append(f"Row {idx+2} ({scb_label}): reasons '{reason}' is not in the allowed list.")
            continue

        sd = _safe_date(row.get("start_date"))
        ed = _safe_date(row.get("end_date"))
        if sd is None or ed is None:
            errors.append(f"Row {idx+2} ({scb_label}): start_date/end_date is invalid or missing.")
            continue
        if sd > ed:
            errors.append(f"Row {idx+2} ({scb_label}): start_date must be <= end_date.")
            continue

        # Integrity: SCB must exist in current SCB OT output set (threshold-filtered)
        if scb_label not in sys_map_dev:
            errors.append(f"Row {idx+2} ({scb_label}): scb_label not present in SCB OT output for the selected filters.")
            continue

        # Integrity: deviation_pct + disconnected_strings must match system values
        up_dev = pd.to_numeric(pd.Series([row.get("deviation_pct")]), errors="coerce").iloc[0]
        up_ds = pd.to_numeric(pd.Series([row.get("disconnected_strings")]), errors="coerce").iloc[0]
        if pd.isna(up_dev):
            errors.append(f"Row {idx+2} ({scb_label}): deviation_pct is invalid.")
            continue
        if pd.isna(up_ds):
            errors.append(f"Row {idx+2} ({scb_label}): disconnected_strings is invalid.")
            continue

        sys_dev = float(sys_map_dev[scb_label])
        sys_ds = int(sys_map_ds[scb_label])

        # Keep strict, but tolerate tiny float round-trip differences from Excel formatting.
        if abs(float(up_dev) - sys_dev) > 1e-6:
            errors.append(
                f"Row {idx+2} ({scb_label}): deviation_pct does not match system value "
                f"(uploaded={float(up_dev):.6f}, system={sys_dev:.6f})."
            )
            continue
        if int(round(float(up_ds))) != sys_ds:
            errors.append(
                f"Row {idx+2} ({scb_label}): disconnected_strings does not match system value "
                f"(uploaded={int(round(float(up_ds)))}, system={sys_ds})."
            )
            continue

        remarks = str(row.get("remarks") or "").strip()

        payloads.append(
            {
                # Required for filtering/reflection and to match existing add_comments schema usage.
                "site_name": str(run.site_name),
                "equipment_names": [scb_label],
                "deviation": float(sys_dev),
                "reasons": [reason],
                "remarks": remarks,
                "start_date": sd.isoformat(),
                "end_date": ed.isoformat(),
                "created_at": _now_utc_iso(),
            }
        )

    return payloads, errors


def render(db_path: str) -> None:
    st.markdown("## SCB Comment")
    st.caption("Excel-driven SCB OT comments. SCB OT is the source of truth; this page only validates + writes comments to Supabase.")

    # Username-based access control (UI/data-filter driven)
    from access_control import allowed_sites_for_user, is_admin, is_restricted_user

    username = st.session_state.get("user_info", {}).get("username")

    # Persist success message for up to 60 seconds (or until tab switch).
    success_until = float(st.session_state.get("scb_comment_success_until", 0.0) or 0.0)
    if success_until and time.time() < success_until:
        st.success("SCB comments submitted successfully.")

    sites = scb_ot.list_sites_from_array_details(db_path)
    if not sites:
        st.info("No sites found in array_details.site_name.")
        return

    allowed_sites = allowed_sites_for_user(username)
    if allowed_sites:
        allowed_l = {str(x).strip().lower() for x in allowed_sites}
        sites = [s for s in sites if str(s).strip().lower() in allowed_l]

    c1, c2, c3, c4 = st.columns([2.6, 1.4, 1.6, 1.6], vertical_alignment="bottom")
    with c1:
        if not is_admin(username) and allowed_sites:
            site_name = st.selectbox("Site Name", options=sites, index=0, disabled=True, key="scb_comment_site_locked")
        else:
            site_name = st.selectbox("Site Name", options=["(select)", *sites], index=0, key="scb_comment_site")
            if site_name == "(select)":
                site_name = None
    with c2:
        threshold = st.number_input(
            "Threshold (%)",
            value=-3.0,
            step=0.5,
            help="Deviation threshold. SCBs with deviation <= threshold will be included in the template.",
            key="scb_comment_threshold",
        )
    with c3:
        from_date = st.date_input("From Date", value=None, key="scb_comment_from")
    with c4:
        to_date = st.date_input("To Date", value=None, key="scb_comment_to")

    # Date availability logic (same backend logic as SCB OT).
    # Streamlit date_input cannot restrict to a set, so we enforce via warning + button disable.
    available_dates: set[date] = set()
    if site_name:
        try:
            table = scb_ot.resolve_site_table_name(db_path, str(site_name))
            cols = scb_ot.get_table_columns(db_path, table)
            scb_cols = [c for c in cols if str(c).upper().startswith("SCB")]
            available_dates = scb_ot._available_dates_for_site_table(db_path, table, tuple(scb_cols))  # type: ignore[attr-defined]
        except Exception:
            available_dates = set()

    if site_name:
        st.caption("Only dates with SCB data (06:00–18:00) are selectable.")
        if available_dates:
            ad_min = min(available_dates)
            ad_max = max(available_dates)
            st.caption(f"Available SCB dates: **{ad_min.isoformat()} → {ad_max.isoformat()}**")
        if from_date and available_dates and from_date not in available_dates:
            st.warning("Selected From Date has no SCB data between 06:00–18:00")
        if to_date and available_dates and to_date not in available_dates:
            st.warning("Selected To Date has no SCB data between 06:00–18:00")

    if site_name and from_date and to_date and from_date > to_date:
        st.warning("From Date must be <= To Date.")
        return

    # Internal state for avoiding unintended reruns/extra compute.
    st.session_state.setdefault("scb_comment_last_run", None)
    st.session_state.setdefault("scb_comment_template_bytes", None)

    st.markdown("---")
    b1, b2 = st.columns([1.2, 1.0], vertical_alignment="bottom")

    with b1:
        gen_clicked = st.button(
            "Generate Excel Template",
            type="primary",
            disabled=not (site_name and from_date and to_date and (not available_dates or (from_date in available_dates and to_date in available_dates))),
            use_container_width=True,
        )

    with b2:
        st.write("")  # alignment spacer

    prog = st.progress(0, text="Idle")

    if gen_clicked:
        try:
            prog.progress(0.0, text="Starting…")
            run = _run_scb_ot_once(
                db_path=db_path,
                site_name=str(site_name),
                from_date=from_date,
                to_date=to_date,
                threshold=float(threshold),
                progress=prog,
            )
            prog.progress(0.75, text="Generating Excel template…")
            xbytes = _build_excel_template_bytes(run=run)
            st.session_state["scb_comment_last_run"] = {
                "site_name": run.site_name,
                "threshold": run.threshold,
                "from_date": run.from_date.isoformat(),
                "to_date": run.to_date.isoformat(),
            }
            st.session_state["scb_comment_template_bytes"] = xbytes
            prog.progress(1.0, text="Template ready")
        except Exception as e:
            prog.progress(0.0, text="Failed")
            st.error(f"Failed to generate template: {e}")

    xbytes = st.session_state.get("scb_comment_template_bytes")
    if xbytes:
        fname = f"SCB_Comment_Template_{st.session_state.get('scb_comment_last_run',{}).get('site_name','SITE')}_{_now_utc_iso().replace(':','-')}.xlsx"
        st.download_button(
            "Download Excel Template",
            data=xbytes,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=False,
        )

    st.markdown("### Upload filled Excel")
    up = st.file_uploader("Upload the filled template (.xlsx)", type=["xlsx"], key="scb_comment_uploader")

    # If Excel is uploaded, auto-derive site + dates so submission works without reselection.
    excel_df = pd.DataFrame()
    excel_site: Optional[str] = None
    excel_from: Optional[date] = None
    excel_to: Optional[date] = None
    if up is not None:
        try:
            excel_df = _read_uploaded_excel(up.getvalue())
            if not excel_df.empty and "site_name" in excel_df.columns:
                excel_site = str(excel_df["site_name"].iloc[0]).strip() if pd.notna(excel_df["site_name"].iloc[0]) else None
            # Derive date bounds from Excel start/end columns
            if not excel_df.empty and "start_date" in excel_df.columns and "end_date" in excel_df.columns:
                sds = [_safe_date(v) for v in excel_df["start_date"].tolist()]
                eds = [_safe_date(v) for v in excel_df["end_date"].tolist()]
                sds2 = [d for d in sds if d is not None]
                eds2 = [d for d in eds if d is not None]
                if sds2:
                    excel_from = min(sds2)
                if eds2:
                    excel_to = max(eds2)
        except Exception:
            excel_df = pd.DataFrame()

    # Effective filters: UI values take precedence if provided; otherwise Excel-derived values.
    eff_site = str(site_name) if site_name else (excel_site or None)
    eff_from = from_date if from_date else excel_from
    eff_to = to_date if to_date else excel_to

    if up is not None:
        if excel_site or excel_from or excel_to:
            st.caption(
                "Excel-derived values (used if the filters above are empty): "
                f"site_name={excel_site or '—'}, from={excel_from.isoformat() if excel_from else '—'}, to={excel_to.isoformat() if excel_to else '—'}"
            )
        if site_name and excel_site and str(site_name).strip().upper() != str(excel_site).strip().upper():
            st.warning("Uploaded Excel site_name does not match the selected Site Name.")

    submit_clicked = st.button(
        "Submit SCB Comments",
        # UX requirement: enabled when a file is uploaded (validation happens on click).
        disabled=(up is None),
        use_container_width=True,
    )

    if submit_clicked:
        # IMPORTANT UX: never `return` from here; View/Edit must still render below.
        if up is None:
            st.error("Upload the filled Excel template before submitting.")
        elif eff_site is None or eff_from is None or eff_to is None:
            st.error("Missing site_name / From Date / To Date. Provide them in the UI or upload a valid Excel with these fields.")
        elif eff_from > eff_to:
            st.error("From Date must be <= To Date.")
        else:
            try:
                prog.progress(0.05, text="Computing SCB OT results for validation…")
                run = _run_scb_ot_once(
                    db_path=db_path,
                    site_name=str(eff_site),
                    from_date=eff_from,
                    to_date=eff_to,
                    threshold=float(threshold),
                    progress=prog,
                )

                prog.progress(0.65, text="Reading uploaded Excel…")
                df_up = _read_uploaded_excel(up.getvalue())

                prog.progress(0.75, text="Validating…")
                payloads, errors = _validate_upload(df_up=df_up, run=run)
                if errors:
                    prog.progress(0.0, text="Validation failed")
                    st.error("Validation failed. Fix the errors and re-upload.")
                    for err in errors[:50]:
                        st.write(f"- {err}")
                    if len(errors) > 50:
                        st.caption(f"(Showing first 50 of {len(errors)} errors)")
                else:
                    prog.progress(0.90, text="Uploading to Supabase…")
                    # Critical requirement: each SCB + reason is a separate row.
                    # We still use bulk insert for efficiency, but each payload is one row.
                    res = add_comments.insert_bulk_comments(payloads)
                    # Ensure SCB OT and View/Edit reflect updates immediately (avoid stale cache).
                    try:
                        add_comments.fetch_comments_live.clear()  # type: ignore[attr-defined]
                    except Exception:
                        pass

                    prog.progress(1.0, text="Done")
                    st.session_state["scb_comment_success_until"] = time.time() + 60.0
                    inserted = int((res or {}).get("inserted", 0))
                    duplicates = int((res or {}).get("duplicates", 0))
                    if inserted > 0:
                        st.success(f"{inserted} comment(s) submitted successfully.")
                    if duplicates > 0:
                        st.warning(f"{duplicates} comment(s) already existed and were skipped.")
                    if inserted == 0 and duplicates == 0:
                        st.info("No comments were submitted.")
            except Exception as e:
                prog.progress(0.0, text="Failed")
                st.error(f"Submission failed: {e}")

    # -----------------------------
    # View / Edit SCB Comments
    # -----------------------------
    st.markdown("### View / Edit SCB Comments")

    # Independent filter block (does NOT depend on Excel/template flow; does NOT call SCB OT).
    ve1, ve2, ve3 = st.columns([2.2, 1.4, 1.4], vertical_alignment="bottom")
    with ve1:
        if not is_admin(username) and allowed_sites:
            ve_site = st.selectbox("Site Name", options=sites, index=0, disabled=True, key="scb_comment_ve_site_locked")
        else:
            ve_site = st.selectbox("Site Name", options=["(select)", *sites], index=0, key="scb_comment_ve_site")
            if ve_site == "(select)":
                ve_site = None
    with ve2:
        ve_from = st.date_input("From Date", value=None, key="scb_comment_ve_from")
    with ve3:
        ve_to = st.date_input("To Date", value=None, key="scb_comment_ve_to")

    if ve_site and ve_from and ve_to and ve_from > ve_to:
        st.warning("From Date must be <= To Date.")
        return
    if not ve_site or not ve_from or not ve_to:
        st.caption("Select Site Name + date range to view/edit SCB comments.")
        return

    # Defensive guard (safety net)
    if is_restricted_user(username) and str(ve_site).strip().lower() != str(username).strip().lower():
        st.error("Unauthorized site access")
        st.stop()

    prog3 = st.progress(0, text="Fetching SCB comments…")
    try:
        base = add_comments.fetch_comments_live(site_name=str(ve_site), start_date=ve_from, end_date=ve_to, limit=2000)
    except Exception as e:
        prog3.progress(0.0, text="Failed")
        st.error(f"Failed to fetch SCB comments: {e}")
        return
    prog3.progress(0.3, text="Filtering & preparing…")

    if base is None or base.empty:
        prog3.progress(1.0, text="Ready")
        st.info("No SCB comments found for the selected site and date range.")
        return

    tmp = base.copy()
    # Normalize equipment_names to a single SCB label string (read-only); keep column name equipment_names.
    if "equipment_names" in tmp.columns:
        tmp = tmp.explode("equipment_names")
        tmp["equipment_names"] = tmp["equipment_names"].astype(str)
    else:
        tmp["equipment_names"] = ""
    tmp["equipment_names"] = tmp["equipment_names"].astype(str)
    tmp["scb_label"] = tmp["equipment_names"]
    tmp = tmp[tmp["equipment_names"].str.contains("SCB", case=False, na=False)].copy()

    # Overlap rule is already enforced in fetch_comments_live via start/end filters.
    # Filters: IS / INV / SCB (string match on scb_label)
    f_is, f_inv, f_scb = st.columns([1.2, 1.2, 1.2], vertical_alignment="bottom")
    with f_is:
        is_filter = st.text_input("Filter IS (e.g. IS8)", value="", key="scb_comment_filter_is")
    with f_inv:
        inv_filter = st.text_input("Filter INV (e.g. INV1)", value="", key="scb_comment_filter_inv")
    with f_scb:
        scb_filter = st.text_input("Filter SCB (e.g. SCB12)", value="", key="scb_comment_filter_scb")

    view = tmp.copy()
    if is_filter.strip():
        view = view[view["scb_label"].str.contains(is_filter.strip(), case=False, na=False)]
    if inv_filter.strip():
        view = view[view["scb_label"].str.contains(inv_filter.strip(), case=False, na=False)]
    if scb_filter.strip():
        view = view[view["scb_label"].str.contains(scb_filter.strip(), case=False, na=False)]

    # Clean display fields
    if "reasons" in view.columns:
        view["reasons"] = view["reasons"].apply(lambda x: ", ".join([str(v) for v in x]) if isinstance(x, list) else ("" if x is None else str(x)))
    if "created_at" in view.columns:
        view["created_at"] = pd.to_datetime(view["created_at"], errors="coerce")
    if "start_date" in view.columns:
        view["start_date"] = pd.to_datetime(view["start_date"], errors="coerce").dt.date.astype(str)
    if "end_date" in view.columns:
        view["end_date"] = pd.to_datetime(view["end_date"], errors="coerce").dt.date.astype(str)

    # Display columns with editable enforcement:
    # - Read-only: site_name, deviation, equipment_names
    # - Editable via form: start_date, end_date, reasons, remarks
    # We keep id for updates.
    show_cols = ["id", "site_name", "deviation", "equipment_names", "start_date", "end_date", "reasons", "remarks", "created_at", "created_by"]
    present = [c for c in show_cols if c in view.columns]
    view_disp = view[present].copy()

    prog3.progress(1.0, text="Ready")

    # Selection + edit form
    selected_row: Optional[dict[str, Any]] = None
    if getattr(add_comments, "_HAS_AGGRID", False) and getattr(add_comments, "AgGrid", None) is not None:
        # Local AgGrid config for selection (Add Comments table is read-only; here we need row selection).
        from st_aggrid import AgGrid  # type: ignore
        from st_aggrid.grid_options_builder import GridOptionsBuilder  # type: ignore
        from st_aggrid.shared import GridUpdateMode  # type: ignore

        # Keep id in the grid for stable selection, but hide it.
        gb = GridOptionsBuilder.from_dataframe(view_disp)
        gb.configure_default_column(sortable=True, filter=True, resizable=True, wrapText=True, autoHeight=True)
        if "id" in view_disp.columns:
            gb.configure_column("id", hide=True)
        gb.configure_selection("single", use_checkbox=True)
        grid_options = gb.build()

        grid = AgGrid(
            view_disp,
            gridOptions=grid_options,
            theme="balham",
            height=360,
            fit_columns_on_grid_load=True,
            allow_unsafe_jscode=False,
            enable_enterprise_modules=False,
            update_mode=GridUpdateMode.SELECTION_CHANGED,
            key="scb_comment_aggrid",
        )
        # AgGrid return type differs by version. Avoid boolean checks on DataFrames (ambiguous truth value).
        selected_rows: list[dict[str, Any]] = []
        if isinstance(grid, dict):
            selected_rows = grid.get("selected_rows") or []
        else:
            selected_rows = getattr(grid, "selected_rows", None) or []
        sel = selected_rows
        if sel:
            s = sel[0]
            m = view_disp
            if "id" in view_disp.columns and s.get("id") is not None:
                m = m[m["id"] == s.get("id")]
            if not m.empty:
                selected_row = m.iloc[0].to_dict()
    else:
        st.dataframe(view_disp.drop(columns=["id"], errors="ignore"), width="stretch", hide_index=True)
        st.caption("Tip: install `st-aggrid` to enable row selection/editing from the table UI.")

    if selected_row:
        st.markdown("#### Edit selected comment")
        cid = selected_row.get("id")
        if cid is None:
            st.error("Selected comment does not have an id; cannot update.")
            return

        # Read-only context
        st.caption(
            f"**Site:** {selected_row.get('site_name','')}  •  "
            f"**Deviation:** {selected_row.get('deviation','')}  •  "
            f"**Equipment:** {selected_row.get('equipment_names','')}"
        )

        c1e, c2e = st.columns([1.4, 1.0], vertical_alignment="bottom")
        with c1e:
            # Preselect existing reason when possible
            existing_reason = None
            r0 = selected_row.get("reasons")
            if isinstance(r0, list) and r0:
                existing_reason = str(r0[0])
            elif isinstance(r0, str) and r0.strip():
                existing_reason = r0.split(",")[0].strip()
            idx = add_comments.REASONS.index(existing_reason) if (existing_reason in add_comments.REASONS) else 0
            reason_new = st.selectbox("Reason", options=add_comments.REASONS, index=idx, key="scb_comment_edit_reason")
        with c2e:
            st.caption(f"SCB: **{selected_row.get('equipment_names','')}**")

        sd_new = st.date_input("Start Date", value=_safe_date(selected_row.get("start_date")) or ve_from, key="scb_comment_edit_start")
        ed_new = st.date_input("End Date", value=_safe_date(selected_row.get("end_date")) or ve_to, key="scb_comment_edit_end")
        remarks_new = st.text_area("Remarks", value=str(selected_row.get("remarks") or ""), key="scb_comment_edit_remarks")

        upd = st.button("Update comment", type="primary", use_container_width=True)
        if upd:
            if sd_new > ed_new:
                st.error("Start Date must be <= End Date.")
                return
            prog_u = st.progress(0, text="Updating comment…")
            prog_u.progress(0.7, text="Updating comment…")
            try:
                payload = {
                    "reasons": [str(reason_new)],
                    "remarks": str(remarks_new or ""),
                    "start_date": sd_new.isoformat(),
                    "end_date": ed_new.isoformat(),
                }
                add_comments.update_comment(cid, payload)
                # Force immediate reflection (avoid cached fetch returning stale data).
                try:
                    add_comments.fetch_comments_live.clear()  # type: ignore[attr-defined]
                except Exception:
                    pass
                prog_u.progress(1.0, text="Update successful")
                st.success("Update successful (will reflect in SCB OT automatically).")
            except Exception as e:
                prog_u.progress(0.0, text="Failed")
                st.error(f"Update failed: {e}")

    # -----------------------------
    # Bulk Edit (optional) — edit multiple rows, then submit changes with progress
    # -----------------------------
    with st.expander("Bulk edit (edit multiple rows then submit)", expanded=False):
        if view_disp is None or view_disp.empty:
            st.info("No rows available for bulk edit.")
        else:
            edit_df = view_disp.copy()
            # Enforce editability: only these columns are editable.
            editable_cols = {"start_date", "end_date", "reasons", "remarks"}
            disabled_cols = [c for c in edit_df.columns if c not in editable_cols]

            edited = st.data_editor(
                edit_df,
                hide_index=True,
                disabled=disabled_cols,
                use_container_width=True,
                key="scb_comment_bulk_editor",
            )

            submit_changes = st.button("Submit Changes", type="primary", key="scb_comment_bulk_submit")
            if submit_changes:
                if edited is None or edited.empty:
                    st.info("No rows to update.")
                else:
                    # Compute changed rows by id
                    base_by_id = {r["id"]: r for r in edit_df.to_dict("records") if r.get("id") is not None}
                    upd_rows: list[dict[str, Any]] = []
                    errs: list[str] = []

                    for r in edited.to_dict("records"):
                        cid = r.get("id")
                        if cid is None or cid not in base_by_id:
                            continue
                        before = base_by_id[cid]
                        # Detect changes only in allowed fields
                        changed = any(
                            str(r.get(k, "")) != str(before.get(k, ""))
                            for k in ["start_date", "end_date", "reasons", "remarks"]
                        )
                        if not changed:
                            continue

                        # Validate dates
                        sd = _safe_date(r.get("start_date"))
                        ed = _safe_date(r.get("end_date"))
                        if sd is None or ed is None:
                            errs.append(f"id={cid}: invalid start_date/end_date")
                            continue
                        if sd > ed:
                            errs.append(f"id={cid}: start_date must be <= end_date")
                            continue

                        # Validate reason against allowed list (expect single reason in text)
                        reason_raw = str(r.get("reasons") or "").strip()
                        # When displayed as "a, b", take the first token as the intended reason.
                        reason_one = reason_raw.split(",")[0].strip() if reason_raw else ""
                        if not reason_one or reason_one not in add_comments.REASONS:
                            errs.append(f"id={cid}: invalid reason '{reason_one}'")
                            continue

                        upd_rows.append(
                            {
                                "id": cid,
                                "payload": {
                                    "start_date": sd.isoformat(),
                                    "end_date": ed.isoformat(),
                                    "reasons": [reason_one],
                                    "remarks": str(r.get("remarks") or ""),
                                },
                            }
                        )

                    if errs:
                        st.error("Cannot submit changes due to validation errors:")
                        for e in errs[:50]:
                            st.write(f"- {e}")
                        if len(errs) > 50:
                            st.caption(f"(Showing first 50 of {len(errs)} errors)")
                    elif not upd_rows:
                        st.info("No changes detected.")
                    else:
                        p = st.progress(0, text="Updating comments…")
                        total = len(upd_rows)
                        ok = 0
                        for i, item in enumerate(upd_rows, start=1):
                            try:
                                add_comments.update_comment(item["id"], item["payload"])
                                ok += 1
                            except Exception as e:
                                st.error(f"Update failed for id={item['id']}: {e}")
                            p.progress(i / total, text=f"Updating comments… ({i}/{total})")

                        try:
                            add_comments.fetch_comments_live.clear()  # type: ignore[attr-defined]
                        except Exception:
                            pass
                        st.success(f"Updated {ok}/{total} comment(s).")


